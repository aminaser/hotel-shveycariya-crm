"""Excel builders for «Счёт-фактура» and «Счёт на оплату».

Layouts replicate the hotel's existing xls documents provided by the owner.
«Счёт-фактура» = tax invoice (Лист1 of «счет фактура (2).xls»).
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.schemas.act import ActDocument
from app.services.act_excel import _amount_to_words_ru

EXECUTOR_SHORT = 'ИП ТРК "Швейцария" Бектурова Ж.К.'
EXECUTOR_LONG = 'ИП Торгово-развлекательный комплекс "Швейцария" Бектурова Ж.К.'
EXECUTOR_ADDRESS = "Алматинская область, г. Текели, ул. Ы. Алтынсарина, 20"
EXECUTOR_BIN = "571031400540"
EXECUTOR_IBAN = "KZ616010311000181781"
EXECUTOR_BANK = 'АО "Народный Банк Казахстана"'
EXECUTOR_BIK = "HSBKKZKX"
EXECUTOR_KBE = "19"
PAYMENT_CODE = "872"
EXECUTOR_DIRECTOR = "Бектурова Ж.К."

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _fmt_date(value: date) -> str:
    return f"{value.day:02d}.{value.month:02d}.{value.year}г."


def _cell(ws, row, col, value="", *, align=LEFT, font=None, border=False, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align
    if font:
        cell.font = font
    if border:
        cell.border = BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def _merge(ws, ref, value="", *, align=LEFT, font=None, border=False):
    ws.merge_cells(ref)
    start, end = ref.split(":")

    def parse(coord: str) -> tuple[int, int]:
        letters = "".join(c for c in coord if c.isalpha())
        row = int("".join(c for c in coord if c.isdigit()))
        col = 0
        for ch in letters:
            col = col * 26 + (ord(ch) - 64)
        return row, col

    r1, c1 = parse(start)
    r2, c2 = parse(end)
    if border:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = BORDER
    return _cell(ws, r1, c1, value, align=align, font=font, border=border)


def _customer_text(document: ActDocument) -> str:
    text = document.customer.name
    if document.customer.address:
        text += f", {document.customer.address}"
    if document.customer.identifier:
        text += f", {document.customer.identifier_label}: {document.customer.identifier}"
    if document.customer.iban:
        text += f", ИИК: {document.customer.iban}"
    return text


def _amount_words(document: ActDocument) -> str:
    words = document.total_amount_words or ""
    if "тиын" not in words:
        words = _amount_to_words_ru(document.total_amount)
    return words


def _build_tax_invoice_sheet(ws, document: ActDocument) -> None:
    """Classic Kazakhstan tax invoice — «счет фактура (2).xls» Лист1."""
    widths = {
        "A": 8,
        "B": 36,
        "C": 8,
        "D": 10,
        "E": 12,
        "F": 14,
        "G": 10,
        "H": 10,
        "I": 14,
        "J": 10,
        "K": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    small = Font(size=8)
    normal = Font(size=9)
    bold = Font(size=9, bold=True)
    title = Font(size=12, bold=True)

    _cell(ws, 4, 4, "Счет-фактура", align=RIGHT, font=title)
    _cell(ws, 4, 6, document.act_number, align=CENTER, font=title)
    _cell(ws, 4, 8, _fmt_date(document.act_date), align=LEFT, font=title)

    header_lines = [
        f"Поставщик: {EXECUTOR_LONG}",
        f"БИН и адрес поставщика: ИИН {EXECUTOR_BIN}  {EXECUTOR_ADDRESS}",
        f"Реквизиты поставщика:  {EXECUTOR_IBAN} в {EXECUTOR_BANK},  БИК {EXECUTOR_BIK}",
        "Договор (контракт) на поставку товаров (работ, услуг): "
        + (document.contract_number or ""),
        "Условия оплаты по договору (контракту): б/нал. расчет",
        "Поставка товаров (работ, услуг) осуществлена по доверенности: Без доверенности",
        "Способ отправления: ",
        "Товарно-транспортная накладная: ",
        f"Грузоотправитель: {EXECUTOR_LONG} {EXECUTOR_ADDRESS}",
        "",
        f"Покупатель:  {document.customer.name}",
        f"РНН и адрес покупателя :  {document.customer.identifier_label}:   "
        f"{document.customer.identifier}"
        + (f",   {document.customer.address}" if document.customer.address else ""),
        f"IBAN покупателя:  ИИК:   {document.customer.iban}"
        if document.customer.iban
        else "IBAN покупателя:  ИИК:   ",
    ]
    row = 5
    for line in header_lines:
        _merge(ws, f"A{row}:K{row}", line, align=LEFT, font=normal)
        row += 1
    row += 1

    head = row
    _merge(ws, f"A{head}:A{head + 1}", "№ п/п", align=CENTER, font=small, border=True)
    _merge(
        ws,
        f"B{head}:B{head + 1}",
        "Наименование товаров (работ, услуг)",
        align=CENTER,
        font=small,
        border=True,
    )
    _merge(ws, f"C{head}:C{head + 1}", "Ед. изм.", align=CENTER, font=small, border=True)
    _merge(ws, f"D{head}:D{head + 1}", "Кол-во (объем)", align=CENTER, font=small, border=True)
    _merge(ws, f"E{head}:E{head + 1}", "Цена тенге", align=CENTER, font=small, border=True)
    _merge(
        ws,
        f"F{head}:F{head + 1}",
        "Стоимость товаров (работ, услуг) без НДС",
        align=CENTER,
        font=small,
        border=True,
    )
    _merge(ws, f"G{head}:H{head}", "НДС", align=CENTER, font=small, border=True)
    _merge(
        ws,
        f"I{head}:I{head + 1}",
        "Всего стоимость реализации",
        align=CENTER,
        font=small,
        border=True,
    )
    _merge(ws, f"J{head}:K{head}", "Акциз", align=CENTER, font=small, border=True)
    _cell(ws, head + 1, 7, "Ставка", align=CENTER, font=small, border=True)
    _cell(ws, head + 1, 8, "Сумма", align=CENTER, font=small, border=True)
    _cell(ws, head + 1, 10, "Ставка", align=CENTER, font=small, border=True)
    _cell(ws, head + 1, 11, "Сумма", align=CENTER, font=small, border=True)

    numbering = head + 2
    for idx, num in enumerate(["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"], start=1):
        _cell(ws, numbering, idx, num, align=CENTER, font=small, border=True)

    row = numbering + 1
    for item in document.line_items:
        _cell(ws, row, 1, item.line_no, align=CENTER, font=normal, border=True)
        _cell(ws, row, 2, item.description, align=LEFT, font=normal, border=True)
        _cell(ws, row, 3, item.unit, align=CENTER, font=normal, border=True)
        _cell(
            ws,
            row,
            4,
            float(item.quantity),
            align=RIGHT,
            font=normal,
            border=True,
            number_format="#,##0.##",
        )
        _cell(
            ws,
            row,
            5,
            float(item.unit_price),
            align=RIGHT,
            font=normal,
            border=True,
            number_format="#,##0.##",
        )
        _cell(
            ws,
            row,
            6,
            float(item.amount),
            align=RIGHT,
            font=normal,
            border=True,
            number_format="#,##0",
        )
        _cell(ws, row, 7, "Без НДС", align=CENTER, font=normal, border=True)
        _cell(ws, row, 8, "", align=RIGHT, font=normal, border=True)
        _cell(
            ws,
            row,
            9,
            float(item.amount),
            align=RIGHT,
            font=normal,
            border=True,
            number_format="#,##0",
        )
        _cell(ws, row, 10, "", align=CENTER, font=normal, border=True)
        _cell(ws, row, 11, "", align=RIGHT, font=normal, border=True)
        row += 1

    _cell(ws, row, 1, "Всего по счету", align=LEFT, font=bold, border=True)
    _merge(ws, f"B{row}:H{row}", f" {_amount_words(document)} ", align=LEFT, font=bold, border=True)
    _cell(
        ws,
        row,
        9,
        float(document.total_amount),
        align=RIGHT,
        font=bold,
        border=True,
        number_format="#,##0",
    )
    _cell(ws, row, 10, "", border=True)
    _cell(ws, row, 11, "", border=True)
    row += 2

    _merge(ws, f"A{row}:F{row}", f"Руководитель:  {EXECUTOR_DIRECTOR}", align=LEFT, font=normal)
    _merge(ws, f"H{row}:K{row}", f"ВЫДАЛ {EXECUTOR_DIRECTOR}", align=LEFT, font=normal)
    row += 1
    _merge(ws, f"A{row}:F{row}", "(Ф.И.О., подпись)", align=CENTER, font=small)
    _merge(ws, f"H{row}:K{row}", "(должность)", align=CENTER, font=small)
    row += 1
    _merge(ws, f"A{row}:F{row}", "Главный бухгалтер  не предусмотрен", align=LEFT, font=normal)
    _merge(ws, f"H{row}:K{row}", "Бухгалтер ", align=LEFT, font=normal)
    row += 1
    _merge(ws, f"A{row}:F{row}", "(Ф.И.О., подпись)", align=CENTER, font=small)
    _merge(ws, f"H{row}:K{row}", "(Ф.И.О., подпись)", align=CENTER, font=small)
    row += 2
    _merge(
        ws,
        f"A{row}:K{row}",
        "Примечание: Без печати не действительно. Оригинал (первый экземпляр) - покупателю. "
        "Копия (второй экземпляр) - поставщику.",
        align=LEFT,
        font=small,
    )

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def build_invoice_workbook(document: ActDocument) -> BytesIO:
    """Счёт-фактура — только налоговая форма, без листа акта."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Счет-фактура"
    _build_tax_invoice_sheet(ws, document)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_payment_invoice_workbook(document: ActDocument) -> BytesIO:
    """Счёт на оплату — layout based on «счет на оплату.xls», TDSheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Счет на оплату"

    widths = {"A": 5, "B": 8, "C": 44, "D": 9, "E": 7, "F": 12, "G": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    small = Font(size=8)
    normal = Font(size=9)
    bold = Font(size=9, bold=True)
    title = Font(size=12, bold=True)

    _merge(
        ws,
        "A1:G3",
        "Внимание! Оплата данного счета означает согласие с условиями поставки товара. "
        "Уведомление об оплате обязательно, в противном случае не гарантируется наличие "
        "товара на складе. Товар отпускается по факту прихода денег на р/с Поставщика, "
        "самовывозом, при наличии доверенности и документов, удостоверяющих личность.",
        align=LEFT_TOP,
        font=small,
    )

    row = 5
    _merge(ws, f"A{row}:G{row}", "Образец платежного поручения", align=CENTER, font=bold)
    row += 1
    _merge(ws, f"A{row}:D{row}", f"Бенефициар: {EXECUTOR_SHORT}", align=LEFT, font=normal, border=True)
    _merge(ws, f"E{row}:F{row}", "ИИК", align=CENTER, font=small, border=True)
    _cell(ws, row, 7, "Кбе", align=CENTER, font=small, border=True)
    row += 1
    _merge(ws, f"A{row}:D{row}", f"БИН: {EXECUTOR_BIN}", align=LEFT, font=normal, border=True)
    _merge(ws, f"E{row}:F{row}", EXECUTOR_IBAN, align=CENTER, font=normal, border=True)
    _cell(ws, row, 7, EXECUTOR_KBE, align=CENTER, font=normal, border=True)
    row += 1
    _merge(ws, f"A{row}:D{row}", f"Банк бенефициара: {EXECUTOR_BANK}", align=LEFT, font=normal, border=True)
    _merge(ws, f"E{row}:F{row}", "БИК", align=CENTER, font=small, border=True)
    _cell(ws, row, 7, "Код назначения платежа", align=CENTER, font=small, border=True)
    row += 1
    _merge(ws, f"A{row}:D{row}", "", border=True)
    _merge(ws, f"E{row}:F{row}", EXECUTOR_BIK, align=CENTER, font=normal, border=True)
    _cell(ws, row, 7, PAYMENT_CODE, align=CENTER, font=normal, border=True)
    row += 2

    _merge(
        ws,
        f"A{row}:G{row}",
        f"Счет на оплату №  {document.act_number}  от  {_fmt_date(document.act_date)}",
        align=Alignment(horizontal="center"),
        font=title,
    )
    row += 2

    _merge(
        ws,
        f"A{row}:G{row}",
        f"Поставщик: {EXECUTOR_SHORT}, {EXECUTOR_ADDRESS}, БИН: {EXECUTOR_BIN}",
        align=LEFT,
        font=normal,
    )
    row += 1
    _merge(ws, f"A{row}:G{row}", f"Покупатель: {_customer_text(document)}", align=LEFT, font=normal)
    row += 2

    _cell(ws, row, 1, "№", align=CENTER, font=bold, border=True)
    _cell(ws, row, 2, "Код", align=CENTER, font=bold, border=True)
    _cell(ws, row, 3, "Наименование", align=CENTER, font=bold, border=True)
    _cell(ws, row, 4, "Кол-во", align=CENTER, font=bold, border=True)
    _cell(ws, row, 5, "Ед.", align=CENTER, font=bold, border=True)
    _cell(ws, row, 6, "Цена", align=CENTER, font=bold, border=True)
    _cell(ws, row, 7, "Сумма", align=CENTER, font=bold, border=True)
    row += 1

    for item in document.line_items:
        description = item.description
        if item.service_date:
            description = f"{description} ({item.service_date})"
        _cell(ws, row, 1, item.line_no, align=CENTER, font=normal, border=True)
        _cell(ws, row, 2, "", border=True)
        _cell(ws, row, 3, description, align=LEFT, font=normal, border=True)
        _cell(ws, row, 4, float(item.quantity), align=RIGHT, font=normal, border=True, number_format="#,##0.##")
        _cell(ws, row, 5, item.unit, align=CENTER, font=normal, border=True)
        _cell(ws, row, 6, float(item.unit_price), align=RIGHT, font=normal, border=True, number_format="#,##0.##")
        _cell(ws, row, 7, float(item.amount), align=RIGHT, font=normal, border=True, number_format="#,##0")
        row += 1

    _merge(ws, f"A{row}:F{row}", "Итого:", align=RIGHT, font=bold, border=True)
    _cell(ws, row, 7, float(document.total_amount), align=RIGHT, font=bold, border=True, number_format="#,##0")
    row += 1

    _merge(
        ws,
        f"A{row}:G{row}",
        f"Всего наименований: {len(document.line_items)}, на сумму {float(document.total_amount):,.2f} ₸".replace(",", " "),
        align=LEFT,
        font=normal,
    )
    row += 1
    _merge(ws, f"A{row}:G{row}", _amount_words(document), align=LEFT, font=bold)
    row += 3

    _merge(ws, f"A{row}:B{row}", "Исполнитель", align=LEFT, font=normal)
    _merge(ws, f"C{row}:D{row}", "_____________________", align=CENTER, font=normal)
    _merge(ws, f"E{row}:G{row}", EXECUTOR_DIRECTOR, align=LEFT, font=normal)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
