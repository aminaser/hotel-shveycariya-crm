from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.schemas.act import ActDocument

EXECUTOR_ACT = (
    'ИП "Бектурова Ж.К." ТРК "Швейцария", Республика Казахстан, '
    "Алматинская область, г. Текели, ул. Алтынсарина 20  "
    'KZ616010311000181781 в АО "Народный Банк Казахстана", БИК HSBKKZKX'
)
EXECUTOR_FULL = (
    'ИП Торгово-развлекательный комплекс "Швейцария" Бектурова Ж.К., '
    "г.Текели, ул. Ы.Алтынсарина 20 "
    'р/с KZ616010311000181781 в АО "Народный Банк Казахстана", БИК HSBKKZKX'
)
EXECUTOR_BIN_DEFAULT = "571031400540"
EXECUTOR_DIRECTOR = "Бектурова Ж.К."

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _format_act_date(value: date) -> str:
    return f"{value.day:02d}.{value.month:02d}.{value.year}г."


def _amount_to_words_ru(amount: Decimal) -> str:
    n = int(amount)
    if n == 0:
        return "ноль тенге 00 тиын"

    ones_f = [
        "",
        "одна",
        "две",
        "три",
        "четыре",
        "пять",
        "шесть",
        "семь",
        "восемь",
        "девять",
        "десять",
        "одиннадцать",
        "двенадцать",
        "тринадцать",
        "четырнадцать",
        "пятнадцать",
        "шестнадцать",
        "семнадцать",
        "восемнадцать",
        "девятнадцать",
    ]
    ones_m = [
        "",
        "один",
        "два",
        "три",
        "четыре",
        "пять",
        "шесть",
        "семь",
        "восемь",
        "девять",
        "десять",
        "одиннадцать",
        "двенадцать",
        "тринадцать",
        "четырнадцать",
        "пятнадцать",
        "шестнадцать",
        "семнадцать",
        "восемнадцать",
        "девятнадцать",
    ]
    tens = [
        "",
        "",
        "двадцать",
        "тридцать",
        "сорок",
        "пятьдесят",
        "шестьдесят",
        "семьдесят",
        "восемьдесят",
        "девяносто",
    ]
    hundreds = [
        "",
        "сто",
        "двести",
        "триста",
        "четыреста",
        "пятьсот",
        "шестьсот",
        "семьсот",
        "восемьсот",
        "девятьсот",
    ]

    def chunk_to_words(num: int, feminine: bool) -> str:
        ones = ones_f if feminine else ones_m
        parts: list[str] = []
        if num >= 100:
            parts.append(hundreds[num // 100])
            num %= 100
        if num >= 20:
            parts.append(tens[num // 10])
            num %= 10
        if num > 0:
            parts.append(ones[num])
        return " ".join(p for p in parts if p)

    def thousand_form(num: int) -> str:
        if num % 10 == 1 and num % 100 != 11:
            return "тысяча"
        if 2 <= num % 10 <= 4 and not (12 <= num % 100 <= 14):
            return "тысячи"
        return "тысяч"

    def million_form(num: int) -> str:
        if num % 10 == 1 and num % 100 != 11:
            return "миллион"
        if 2 <= num % 10 <= 4 and not (12 <= num % 100 <= 14):
            return "миллиона"
        return "миллионов"

    millions = n // 1_000_000
    thousands = (n % 1_000_000) // 1000
    remainder = n % 1000

    words: list[str] = []
    if millions:
        words.append(chunk_to_words(millions, False))
        words.append(million_form(millions))
    if thousands:
        words.append(chunk_to_words(thousands, True))
        words.append(thousand_form(thousands))
    if remainder or not words:
        words.append(chunk_to_words(remainder, False))

    text = " ".join(words).strip()
    return f"{text[0].upper()}{text[1:]} тенге 00 тиын"


def _cell(
    ws,
    row: int,
    col: int,
    value="",
    *,
    align=LEFT,
    font=None,
    border=False,
    number_format: str | None = None,
):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align
    if font:
        cell.font = font
    if border:
        cell.border = BORDER
    if number_format:
        cell.number_format = number_format
    return cell


def _merge(ws, ref: str, value="", *, align=LEFT, font=None, border=False):
    ws.merge_cells(ref)
    start = ref.split(":")[0]
    col_letters = "".join(c for c in start if c.isalpha())
    row = int("".join(c for c in start if c.isdigit()))
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - 64)
    return _cell(ws, row, col, value, align=align, font=font, border=border)


def _parse_service_date(raw: str) -> date | None:
    text = (raw or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _period_from_lines(document: ActDocument | None) -> tuple[str, str]:
    if not document:
        return "", ""
    dates: list[date] = []
    for item in document.line_items:
        parsed = _parse_service_date(item.service_date)
        if parsed:
            dates.append(parsed)
    if not dates:
        return "", ""
    return _format_act_date(min(dates)), _format_act_date(max(dates))


def fill_hotel_act_sheet(ws, document: ActDocument | None = None) -> None:
    """Hotel working act form from «счет фактура (2).xls» Лист2 (cleaned)."""
    widths = {
        "A": 8,
        "B": 28,
        "C": 16,
        "D": 22,
        "E": 10,
        "F": 12,
        "G": 14,
        "H": 12,
        "I": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    small = Font(size=8)
    normal = Font(size=9)
    bold = Font(size=9, bold=True)
    title = Font(size=11, bold=True)

    customer_id = document.customer.identifier if document else ""
    executor_id = (document.executor.identifier if document else "") or EXECUTOR_BIN_DEFAULT
    customer_line = ""
    customer_iban_line = ""
    if document:
        customer_line = document.customer.name
        if document.customer.address:
            customer_line = f"{customer_line}, {document.customer.address}"
        if document.customer.iban:
            customer_iban_line = f"ИИК {document.customer.iban}"

    _cell(ws, 1, 8, "ИИН/БИН", align=CENTER, font=small, border=True)
    if customer_iban_line:
        _merge(ws, "A2:G2", f"Заказчик  {customer_line}".rstrip(), align=LEFT_TOP, font=normal, border=True)
        _merge(ws, "A3:G3", customer_iban_line, align=LEFT_TOP, font=normal, border=True)
    else:
        _merge(
            ws,
            "A2:G3",
            f"Заказчик  {customer_line}".rstrip(),
            align=LEFT_TOP,
            font=normal,
            border=True,
        )
    _merge(ws, "H2:I3", customer_id, align=CENTER, font=normal, border=True)

    _cell(ws, 5, 8, "ИИН/БИН", align=CENTER, font=small, border=True)
    _merge(ws, "A6:G7", f"Исполнитель {EXECUTOR_ACT}", align=LEFT_TOP, font=normal, border=True)
    _merge(ws, "H6:I7", executor_id, align=CENTER, font=normal, border=True)

    _merge(
        ws,
        "A9:E9",
        "АКТ ВЫПОЛНЕННЫХ РАБОТ (ОКАЗАННЫХ УСЛУГ)",
        align=Alignment(horizontal="center", vertical="center"),
        font=title,
    )

    _cell(ws, 9, 6, "Номер документа", align=CENTER, font=small, border=True)
    _cell(ws, 9, 7, "Дата составления", align=CENTER, font=small, border=True)
    _merge(ws, "H9:I9", "Отчетный период", align=CENTER, font=small, border=True)

    _cell(
        ws,
        10,
        6,
        document.act_number if document else "",
        align=CENTER,
        font=normal,
        border=True,
    )
    _cell(
        ws,
        10,
        7,
        _format_act_date(document.act_date) if document else "",
        align=CENTER,
        font=normal,
        border=True,
    )
    period_from, period_to = _period_from_lines(document)
    _cell(
        ws,
        10,
        8,
        f"с {period_from}" if period_from else "с",
        align=CENTER,
        font=small,
        border=True,
    )
    _cell(
        ws,
        10,
        9,
        f"по {period_to}" if period_to else "по",
        align=CENTER,
        font=small,
        border=True,
    )

    _merge(ws, "A12:A13", "Номер\nпо порядку", align=CENTER, font=small, border=True)
    _merge(ws, "B12:B13", "Наименование работ (услуг)", align=CENTER, font=small, border=True)
    _merge(
        ws,
        "C12:C13",
        "Дата выполнения работ (оказания услуг)",
        align=CENTER,
        font=small,
        border=True,
    )
    _merge(
        ws,
        "D12:D13",
        "Сведения о наличии отчета о маркетинговых исследованиях, "
        "консультационных и прочих услуг (дата, номер, количество страниц)",
        align=CENTER,
        font=small,
        border=True,
    )
    _merge(ws, "E12:E13", "Единица измерения", align=CENTER, font=small, border=True)
    _merge(ws, "F12:H12", "Выполнено работ (оказано услуг)", align=CENTER, font=small, border=True)
    _cell(ws, 13, 6, "количество", align=CENTER, font=small, border=True)
    _cell(ws, 13, 7, "цена за единицу, Без НДС", align=CENTER, font=small, border=True)
    _cell(ws, 13, 8, "Сумма, в тенге", align=CENTER, font=small, border=True)
    _cell(ws, 12, 9, "", border=True)
    _cell(ws, 13, 9, "", border=True)

    # Нумерация граф как в шаблоне: дата без номера, сумма = 7
    col_numbers = {1: 1, 2: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7}
    for col in range(1, 10):
        _cell(ws, 14, col, col_numbers.get(col, ""), align=CENTER, font=small, border=True)

    ws.row_dimensions[12].height = 36
    ws.row_dimensions[13].height = 36

    row = 15
    line_items = list(document.line_items) if document else []
    if line_items:
        for item in line_items:
            _cell(ws, row, 1, item.line_no, align=CENTER, font=normal, border=True)
            _cell(ws, row, 2, item.description, align=LEFT, font=normal, border=True)
            _cell(ws, row, 3, item.service_date or "", align=CENTER, font=normal, border=True)
            _cell(ws, row, 4, "", align=LEFT, font=normal, border=True)
            _cell(ws, row, 5, item.unit, align=CENTER, font=normal, border=True)
            _cell(
                ws,
                row,
                6,
                float(item.quantity),
                align=RIGHT,
                font=normal,
                border=True,
                number_format="#,##0.##",
            )
            _cell(
                ws,
                row,
                7,
                float(item.unit_price),
                align=RIGHT,
                font=normal,
                border=True,
                number_format="#,##0.##",
            )
            _cell(
                ws,
                row,
                8,
                float(item.amount),
                align=RIGHT,
                font=normal,
                border=True,
                number_format="#,##0.00",
            )
            _cell(ws, row, 9, "", border=True)
            row += 1
    else:
        for index in range(5):
            _cell(ws, row, 1, index + 1, align=CENTER, font=normal, border=True)
            for col in range(2, 10):
                _cell(ws, row, col, "", border=True)
            row += 1

    for _ in range(2):
        for col in range(1, 10):
            _cell(ws, row, col, "", border=True)
        row += 1

    total_row = row
    for col in range(1, 5):
        _cell(ws, total_row, col, "", border=True)
    _cell(ws, total_row, 5, "Итого", align=CENTER, font=bold, border=True)
    _cell(ws, total_row, 6, "", border=True)
    _cell(ws, total_row, 7, "", border=True)
    total_value = float(document.total_amount) if document else 0
    _cell(
        ws,
        total_row,
        8,
        total_value if document else "",
        align=RIGHT,
        font=bold,
        border=True,
        number_format="#,##0.00",
    )
    _cell(ws, total_row, 9, "", border=True)
    row = total_row + 2

    _merge(
        ws,
        f"A{row}:I{row}",
        "Сведения об использовании запасов, полученных от заказчика"
        "_______________________________________________________________",
        align=LEFT,
        font=normal,
    )
    row += 1
    _merge(ws, f"A{row}:I{row}", "наименование, количество, стоимость", align=CENTER, font=small)
    row += 2
    _merge(
        ws,
        f"A{row}:I{row}",
        "Приложение: Перечень документации"
        "_______________________________________________________________",
        align=LEFT,
        font=normal,
    )
    row += 3

    _merge(
        ws,
        f"A{row}:D{row}",
        "Сдал  _____________/_______________/___________________",
        align=LEFT,
        font=normal,
    )
    _merge(
        ws,
        f"F{row}:I{row}",
        "Принял __________________/_____________/__________________",
        align=LEFT,
        font=normal,
    )
    row += 1
    _merge(
        ws,
        f"A{row}:D{row}",
        "должность           подпись           расшифровка подписи",
        align=LEFT,
        font=small,
    )
    _merge(
        ws,
        f"F{row}:I{row}",
        "должность             подпись                   расшифровка подписи",
        align=LEFT,
        font=small,
    )
    row += 2
    _cell(ws, row, 1, "М.П.", align=LEFT, font=normal)
    _cell(ws, row, 6, "М.П.", align=LEFT, font=normal)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = f"A1:I{row}"


def build_act_workbook(document: ActDocument | None = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Акт выполненных работ"
    fill_hotel_act_sheet(ws, document)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_act_template_workbook() -> BytesIO:
    return build_act_workbook(None)
